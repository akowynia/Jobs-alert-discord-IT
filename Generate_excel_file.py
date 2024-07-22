
import openpyxl
from class_folder.webhook_send import webhook_send
from configparser import RawConfigParser
class Generate_excel_file:
    def __init__(self) -> None:
         pass

    def create_excel_file(self, data ):

              
        file_path = 'jobs_discord.xlsx'
        headers = ['nr', 'Nazwa strony', 'czas','link','tytuł','region']

        
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Ustawienie nagłówków w pierwszym wierszu
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        for row in data:
            next_row = sheet.max_row + 1
            sheet.cell(row=next_row, column=1, value=row[0])
            sheet.cell(row=next_row, column=2, value=row[1])
            sheet.cell(row=next_row, column=3, value=row[2])
            sheet.cell(row=next_row, column=4, value=row[3])
            sheet.cell(row=next_row, column=5, value=row[4])
            sheet.cell(row=next_row, column=6, value=row[5])
            workbook.save(file_path)

                
            

        workbook.save(file_path)
        print("Excel file created")

        config = RawConfigParser()
        config_path = "configs/files_excel.ini"
        config.read(config_path)
        sections = config.sections()
        try:
            for list_website in sections:
                if list_website != "Template":
                    
                    webhook = webhook_send()
                    
                    webhook.send_message_excel(config[list_website]["username"], config[list_website]["avatar_url_discord"], config[list_website]["url_webhook_discord"])
        except Exception as e:
            print("Error",e)

        
        


        
        


